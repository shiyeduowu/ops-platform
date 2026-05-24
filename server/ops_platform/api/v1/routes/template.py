# -*- coding: utf-8 -*-
"""
模板生成器 API

提供Excel模板解析和数据生成功能
"""

import asyncio
import re
import time
import uuid
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional
from collections import OrderedDict

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from pydantic import BaseModel

from ops_platform.api.deps import get_current_user
from ops_platform.schemas import UserContext

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/template", tags=["模板生成器"])

# 存储目录
UPLOAD_DIR = Path("uploads/templates")
OUTPUT_DIR = Path("outputs/templates")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


# ============================================================================
# 缓存管理（带过期清理）
# ============================================================================

class CacheEntry:
    __slots__ = ('filepath', 'analysis', 'created_at')

    def __init__(self, filepath: str, analysis):
        self.filepath = filepath
        self.analysis = analysis
        self.created_at = time.time()


class ParseCache:
    """带过期清理的缓存"""

    def __init__(self, max_size: int = 100, ttl_seconds: int = 3600):
        self._cache: OrderedDict[str, CacheEntry] = OrderedDict()
        self._max_size = max_size
        self._ttl = ttl_seconds

    def _evict_expired(self):
        now = time.time()
        expired = [
            k for k, v in self._cache.items()
            if now - v.created_at > self._ttl
        ]
        for k in expired:
            del self._cache[k]

    def get(self, key: str) -> Optional[CacheEntry]:
        self._evict_expired()
        entry = self._cache.get(key)
        if entry:
            self._cache.move_to_end(key)
        return entry

    def set(self, key: str, entry: CacheEntry):
        self._evict_expired()
        if len(self._cache) >= self._max_size:
            self._cache.popitem(last=False)
        self._cache[key] = entry

    def __contains__(self, key: str) -> bool:
        self._evict_expired()
        return key in self._cache


parse_cache = ParseCache(max_size=100, ttl_seconds=3600)

# 延迟加载引擎
_engine = None


def get_engine():
    global _engine
    if _engine is None:
        from ops_platform.modules.smart_template import SmartTemplateEngine
        _engine = SmartTemplateEngine()
    return _engine


def sanitize_filename(filename: str) -> str:
    """清理文件名，防止路径遍历"""
    # 只保留字母、数字、下划线、短横线、点
    safe = re.sub(r'[^\w\-\.]', '_', filename)
    # 移除连续的点（防止 ../）
    safe = re.sub(r'\.{2,}', '.', safe)
    return safe


# ============================================================================
# 请求/响应模型
# ============================================================================

class GenerateRequest(BaseModel):
    file_id: str
    count: int = 30
    group_size: int = 0


class ColumnInfo(BaseModel):
    index: int
    header: str
    field_id: Optional[str] = None
    confidence: float = 0
    generator: Optional[str] = None


class RelationshipInfo(BaseModel):
    type: str
    parent_col: int
    child_col: int
    description: str


class UploadResponse(BaseModel):
    success: bool
    file_id: Optional[str] = None
    filename: Optional[str] = None
    headers: Optional[list] = None
    columns: Optional[list] = None
    header_row: Optional[int] = None
    relationships: Optional[list] = None
    sample_data: Optional[list] = None
    error: Optional[str] = None


class GenerateResponse(BaseModel):
    success: bool
    filename: Optional[str] = None
    download_url: Optional[str] = None
    total_rows: Optional[int] = None
    groups: Optional[int] = None
    rows_per_group: Optional[int] = None
    preview: Optional[list] = None
    error: Optional[str] = None


# ============================================================================
# API 路由
# ============================================================================

@router.post("/upload", response_model=UploadResponse)
async def upload_template(
    file: UploadFile = File(...),
    current_user: UserContext = Depends(get_current_user),
):
    """上传并解析Excel模板（需要登录）"""
    if not file.filename:
        raise HTTPException(400, "没有选择文件")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {'.xlsx', '.xls'}:
        raise HTTPException(400, "只支持 .xlsx, .xls 文件")

    try:
        # 保存文件（使用安全的文件名）
        safe_name = sanitize_filename(file.filename)
        file_id = f"{uuid.uuid4().hex[:8]}_{safe_name}"
        filepath = UPLOAD_DIR / file_id

        content = await file.read()
        if len(content) > 10 * 1024 * 1024:  # 10MB 限制
            raise HTTPException(413, "文件过大，请上传小于10MB的文件")

        # 异步写入文件
        await asyncio.to_thread(filepath.write_bytes, content)
        logger.info(f"文件上传成功: {file.filename} (user={current_user.user_id})")

        # 异步解析模板（CPU 密集型操作）
        def _parse_template():
            from openpyxl import load_workbook
            engine = get_engine()
            wb = load_workbook(str(filepath), data_only=True)
            analysis = engine.analyze(wb)

            columns = []
            for col_idx, header in enumerate(analysis.headers):
                col_info = {
                    'index': col_idx,
                    'header': header,
                    'field_id': None,
                    'confidence': 0,
                    'generator': None
                }
                if col_idx in analysis.field_matches:
                    match = analysis.field_matches[col_idx]
                    col_info['field_id'] = match.field_id
                    col_info['confidence'] = round(match.confidence, 2)
                    gen = engine.registry.get_best_generator(match.field_id)
                    if gen:
                        col_info['generator'] = gen.generator_id
                columns.append(col_info)

            relationships = []
            for rel in analysis.relationships:
                relationships.append({
                    'type': rel.rel_type.value,
                    'parent_col': rel.parent_col,
                    'child_col': rel.child_col,
                    'description': rel.description
                })

            sheet = wb.active
            sample_data = []
            for row in range(analysis.header_row + 1, min(analysis.header_row + 6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, len(analysis.headers) + 1):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(str(cell.value) if cell.value else '')
                sample_data.append(row_data)

            return analysis, columns, relationships, sample_data

        analysis, columns, relationships, sample_data = await asyncio.to_thread(_parse_template)

        # 缓存（按租户隔离）
        cache_key = f"{current_user.tenant_id}:{file_id}"
        parse_cache.set(cache_key, CacheEntry(str(filepath), analysis))

        return UploadResponse(
            success=True,
            file_id=file_id,
            filename=file.filename,
            headers=analysis.headers,
            columns=columns,
            header_row=analysis.header_row,
            relationships=relationships,
            sample_data=sample_data
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"上传失败: {e}", exc_info=True)
        raise HTTPException(500, f"上传失败: {str(e)}")


@router.post("/generate", response_model=GenerateResponse)
async def generate_data(
    req: GenerateRequest,
    current_user: UserContext = Depends(get_current_user),
):
    """批量生成数据（需要登录）"""
    cache_key = f"{current_user.tenant_id}:{req.file_id}"
    entry = parse_cache.get(cache_key)
    if not entry:
        raise HTTPException(400, "模板文件已过期，请重新上传")

    if req.count <= 0 or req.count > 10000:
        raise HTTPException(400, "生成数量范围: 1-10000")

    try:
        # 异步生成数据（CPU 密集型操作）
        def _generate():
            from openpyxl import load_workbook
            engine = get_engine()

            wb = load_workbook(entry.filepath, data_only=True)
            plan = engine.plan(entry.analysis, count=req.count, group_size=req.group_size)
            data_rows = engine.generate(plan, count=req.count)

            timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
            output_filename = f"generated_{timestamp}_{req.count}rows.xlsx"
            output_path = str(OUTPUT_DIR / output_filename)

            engine.export(wb, plan, data_rows, output_path)
            return output_filename, data_rows, plan

        output_filename, data_rows, plan = await asyncio.to_thread(_generate)

        logger.info(f"生成成功: {output_filename}, {len(data_rows)}行 (user={current_user.user_id})")

        preview = data_rows[:5] if len(data_rows) > 5 else data_rows

        return GenerateResponse(
            success=True,
            filename=output_filename,
            download_url=f"/api/v1/template/download/{output_filename}",
            total_rows=len(data_rows),
            groups=plan.total_groups,
            rows_per_group=plan.rows_per_group,
            preview=preview
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"生成失败: {e}", exc_info=True)
        raise HTTPException(500, f"生成失败: {str(e)}")


@router.get("/download/{filename}")
async def download_file(
    filename: str,
    current_user: UserContext = Depends(get_current_user),
):
    """下载生成的文件（需要登录）"""
    from fastapi.responses import FileResponse

    # 清理文件名，防止路径遍历
    safe_filename = sanitize_filename(filename)

    # 确保文件在 OUTPUT_DIR 内
    filepath = (OUTPUT_DIR / safe_filename).resolve()
    if not str(filepath).startswith(str(OUTPUT_DIR.resolve())):
        raise HTTPException(400, "无效的文件名")

    if not filepath.exists():
        raise HTTPException(404, "文件不存在")

    return FileResponse(
        str(filepath),
        filename=safe_filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/generators")
async def list_generators(
    current_user: UserContext = Depends(get_current_user),
):
    """获取所有可用的生成器（需要登录）"""
    engine = get_engine()
    generators = []
    for gen in engine.registry.get_all_generators():
        generators.append({
            'id': gen.generator_id,
            'fields': gen.compatible_fields,
            'priority': gen.priority()
        })
    return {'success': True, 'generators': generators}
