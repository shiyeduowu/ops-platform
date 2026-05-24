# -*- coding: utf-8 -*-
"""
智能模板引擎

统一入口，编排所有子系统：
1. 表头检测（多策略投票）
2. 列语义分析（同义词匹配）
3. 关系推断（列间耦合）
4. 数据生成（插件化生成器）
5. Excel导出（样式保留）
"""

from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
from dataclasses import dataclass

from .detector import HeaderDetector, HeaderCandidate
from .analyzer import (
    ColumnProfiler, SemanticFieldMatcher,
    RelationshipInferencer, ColumnRelationship
)
from .generator import GeneratorRegistry
from .generator.base import ColumnProfile, GenerationContext, FieldMatch
from .learner import PatternStore, PatternMatcher, LearnedPattern
from .config.settings import TemplateSettings


@dataclass
class TemplateAnalysis:
    """模板分析结果"""
    header_row: int
    headers: List[str]
    column_profiles: List[ColumnProfile]
    field_matches: Dict[int, FieldMatch]       # {列索引: 匹配结果}
    relationships: List[ColumnRelationship]
    pattern: Optional[LearnedPattern] = None


@dataclass
class GenerationPlan:
    """数据生成计划"""
    header_row: int
    headers: List[str]
    generators: Dict[int, Any]                 # {列索引: 生成器}
    configs: Dict[int, Dict]                   # {列索引: 配置}
    relationships: List[ColumnRelationship]
    rows_per_group: int = 1
    total_groups: int = 1


class SmartTemplateEngine:
    """
    智能模板引擎

    功能：
    1. analyze: 分析模板结构
    2. plan: 生成数据生成计划
    3. generate: 生成填充数据
    4. export: 导出到 Excel
    """

    def __init__(self, settings: TemplateSettings = None):
        self._settings = settings or TemplateSettings()
        self._detector = HeaderDetector(self._settings)
        self._profiler = ColumnProfiler()
        self._matcher = SemanticFieldMatcher(self._settings)
        self._inferencer = RelationshipInferencer()
        self._registry = GeneratorRegistry()
        self._pattern_store = PatternStore()
        self._pattern_matcher = PatternMatcher(self._pattern_store)

    @property
    def registry(self) -> GeneratorRegistry:
        """获取生成器注册表（用于插件注册）"""
        return self._registry

    def analyze(self, workbook) -> TemplateAnalysis:
        """
        分析模板

        Args:
            workbook: openpyxl Workbook 对象

        Returns:
            模板分析结果
        """
        sheet = workbook.active

        # 1. 检测表头行
        header_candidate = self._detector.detect(sheet)
        if not header_candidate:
            raise ValueError("无法检测到表头行")
        header_row = header_candidate.row_index

        # 2. 提取列名
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row, column=col)
            val = str(cell.value).strip() if cell.value else ''
            headers.append(val)

        # 3. 分析每列
        column_profiles = []
        field_matches = {}
        for col_idx, header in enumerate(headers):
            # 获取样本值
            sample_values = self._get_sample_values(sheet, header_row, col_idx)

            # 统计分析
            profile = self._profiler.profile(col_idx, header, sample_values)
            column_profiles.append(profile)

            # 语义匹配
            match = self._matcher.match_best(header, sample_values, profile)
            if match:
                field_matches[col_idx] = match

        # 4. 推断列间关系
        relationships = self._inferencer.infer(sheet, header_row, column_profiles)

        # 5. 尝试匹配已学习的模式
        pattern = self._pattern_matcher.match(headers)

        return TemplateAnalysis(
            header_row=header_row,
            headers=headers,
            column_profiles=column_profiles,
            field_matches=field_matches,
            relationships=relationships,
            pattern=pattern
        )

    def plan(
        self,
        analysis: TemplateAnalysis,
        count: int,
        group_size: int = 0
    ) -> GenerationPlan:
        """
        生成数据生成计划

        Args:
            analysis: 模板分析结果
            count: 需要生成的数据行数
            group_size: 每组人数（0表示自动）

        Returns:
            生成计划
        """
        # 自动计算组信息
        if group_size <= 0:
            # 尝试从关系中推断
            group_size = self._infer_group_size(analysis)

        total_groups = (count + group_size - 1) // group_size if group_size > 0 else 1

        # 为每列选择生成器和配置
        generators = {}
        configs = {}

        for col_idx, profile in enumerate(analysis.column_profiles):
            match = analysis.field_matches.get(col_idx)
            if match:
                field_id = match.field_id
                gen = self._registry.get_best_generator(field_id, profile)
                if gen:
                    generators[col_idx] = gen
                    # 配置生成器
                    sample_values = profile.sample_values
                    config = gen.configure(sample_values, profile)
                    configs[col_idx] = config
            else:
                # 尝试使用默认生成器
                gen = self._registry.get_best_generator('', profile)
                if gen:
                    generators[col_idx] = gen
                    configs[col_idx] = gen.configure([], profile)

        return GenerationPlan(
            header_row=analysis.header_row,
            headers=analysis.headers,
            generators=generators,
            configs=configs,
            relationships=analysis.relationships,
            rows_per_group=group_size,
            total_groups=total_groups
        )

    def generate(
        self,
        plan: GenerationPlan,
        count: int
    ) -> List[List[Any]]:
        """
        生成填充数据

        Args:
            plan: 生成计划
            count: 生成行数

        Returns:
            二维数组，每行是一条记录
        """
        rows = []
        group_size = plan.rows_per_group if plan.rows_per_group > 0 else count

        for row_idx in range(count):
            group_index = row_idx // group_size if group_size > 0 else 0
            seat_index = row_idx % group_size if group_size > 0 else 0

            row = []
            for col_idx in range(len(plan.headers)):
                gen = plan.generators.get(col_idx)
                config = plan.configs.get(col_idx, {})

                if gen:
                    context = GenerationContext(
                        row_index=row_idx,
                        group_index=group_index,
                        seat_index=seat_index,
                        column_config=config,
                        total_rows=count,
                        total_groups=plan.total_groups,
                        rows_per_group=group_size
                    )
                    value = gen.generate(context)
                    row.append(value)
                else:
                    row.append('')

            rows.append(row)

        return rows

    def export(
        self,
        workbook,
        plan: GenerationPlan,
        data: List[List[Any]],
        output_path: str
    ):
        """
        导出到 Excel

        保留原模板样式，填充生成的数据。

        Args:
            workbook: 原始模板 workbook
            plan: 生成计划
            data: 生成的数据
            output_path: 输出文件路径
        """
        from openpyxl import Workbook
        from copy import copy

        sheet = workbook.active
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = sheet.title

        # 复制表头（包括样式）
        for col_idx, header in enumerate(plan.headers):
            src_cell = sheet.cell(row=plan.header_row, column=col_idx + 1)
            dst_cell = new_sheet.cell(row=1, column=col_idx + 1, value=src_cell.value)

            # 复制样式
            if src_cell.font:
                dst_cell.font = copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy(src_cell.fill)
            if src_cell.alignment:
                dst_cell.alignment = copy(src_cell.alignment)
            if src_cell.border:
                dst_cell.border = copy(src_cell.border)
            if src_cell.number_format:
                dst_cell.number_format = src_cell.number_format

        # 填充数据
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                new_sheet.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        # 复制列宽
        for col_idx in range(1, len(plan.headers) + 1):
            col_letter = new_sheet.cell(row=1, column=col_idx).column_letter
            if col_letter in sheet.column_dimensions:
                new_sheet.column_dimensions[col_letter].width = \
                    sheet.column_dimensions[col_letter].width

        # 保存
        new_wb.save(output_path)

    def analyze_and_generate(
        self,
        workbook,
        count: int,
        group_size: int = 0
    ) -> Tuple[TemplateAnalysis, GenerationPlan, List[List[Any]]]:
        """
        一键分析并生成

        Args:
            workbook: openpyxl Workbook 对象
            count: 生成行数
            group_size: 每组人数

        Returns:
            (分析结果, 生成计划, 生成数据)
        """
        analysis = self.analyze(workbook)
        plan = self.plan(analysis, count, group_size)
        data = self.generate(plan, count)
        return analysis, plan, data

    def _get_sample_values(
        self,
        sheet,
        header_row: int,
        col_idx: int,
        max_samples: int = 10
    ) -> List[Any]:
        """获取列的样本值"""
        values = []
        for row in range(header_row + 1, min(header_row + max_samples + 1, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col_idx + 1)
            if cell.value is not None:
                values.append(cell.value)
        return values

    def _infer_group_size(self, analysis: TemplateAnalysis) -> int:
        """从关系中推断每组人数"""
        for rel in analysis.relationships:
            if rel.rel_type.value == 'group_seat':
                # 找座位号列的循环长度
                seat_col = rel.child_col
                profile = analysis.column_profiles[seat_col]
                if profile.cycle_length:
                    return profile.cycle_length

        # 默认不分组
        return 0

    def save_pattern(self, analysis: TemplateAnalysis, user_mappings: Dict[int, str]):
        """
        保存用户修正的映射

        Args:
            analysis: 模板分析结果
            user_mappings: 用户修正的 {列索引: 字段ID}
        """
        # 合并自动匹配和用户修正
        column_mappings = {}
        for col_idx, match in analysis.field_matches.items():
            column_mappings[str(col_idx)] = match.field_id
        for col_idx, field_id in user_mappings.items():
            column_mappings[str(col_idx)] = field_id

        # 计算签名
        signature = self._pattern_store.compute_signature(analysis.headers)

        # 创建模式
        pattern = LearnedPattern(
            pattern_id=f"pattern_{signature}",
            template_signature=signature,
            header_row=analysis.header_row,
            column_mappings=column_mappings,
            column_configs={}
        )

        self._pattern_store.add_pattern(pattern)
